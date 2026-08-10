"""Três pedidos do módulo Cadastro e dados:
(a) administrador cadastra entidade gestora + usuário responsável por ela;
(b) modelo de planilha pra ajudar quem vai importar;
(c) administrador e gestores (PAPEIS_GESTAO) cadastram entidade nova
    direto pra uma CPL, sem precisar da API crua."""

import csv
import io

import openpyxl
from conftest import criar_usuario_com_papel, login_como

from app.models.enums import Papel


def _criar_cpl(admin_client, sigla="CPL-GESTORA-01") -> str:
    resposta = admin_client.post("/api/cpls", json={"nome": "CPL de Teste", "sigla": sigla})
    assert resposta.status_code == 201
    return resposta.json()["id"]


# --- B: modelo de planilha ----------------------------------------------------


def test_modelo_planilha_web_xlsx_tem_so_cabecalho(admin_client):
    resposta = admin_client.get("/painel/cadastro/modelo-planilha?formato=xlsx")
    assert resposta.status_code == 200
    pasta = openpyxl.load_workbook(io.BytesIO(resposta.content))
    planilha = pasta.active
    assert planilha.max_row == 1
    assert planilha.cell(1, 1).value == "razao_social"


def test_modelo_planilha_api_csv_tem_so_cabecalho(admin_client):
    resposta = admin_client.get("/api/cadastro/modelo-planilha?formato=csv")
    assert resposta.status_code == 200
    linhas = list(csv.reader(io.StringIO(resposta.content.decode("utf-8-sig"))))
    assert len(linhas) == 1
    assert linhas[0][0] == "razao_social"


def test_modelo_planilha_exige_papel_gestao(client_sem_papel):
    resposta = client_sem_papel.get("/painel/cadastro/modelo-planilha?formato=xlsx")
    assert resposta.status_code == 403


# --- C: cadastrar entidade nova direto pra uma CPL -----------------------------


def test_criar_entidade_para_cpl_cria_e_vincula(admin_client):
    cpl_id = _criar_cpl(admin_client, sigla="CPL-GESTORA-02")
    resposta = admin_client.post(
        f"/painel/cadastro/cpls/{cpl_id}/entidades",
        data={"tipo": "empresa", "razao_social": "Nova Empresa Ltda", "municipio": "Atibaia", "uf": "SP"},
        follow_redirects=False,
    )
    assert resposta.status_code == 303

    vinculos = admin_client.get(f"/api/cpls/{cpl_id}/entidades").json()
    assert any(v["entidade"]["razao_social"] == "Nova Empresa Ltda" for v in vinculos)


def test_criar_entidade_para_cpl_valida_cnpj(admin_client):
    cpl_id = _criar_cpl(admin_client, sigla="CPL-GESTORA-03")
    resposta = admin_client.post(
        f"/painel/cadastro/cpls/{cpl_id}/entidades",
        data={"tipo": "empresa", "razao_social": "CNPJ Invalido Ltda", "cnpj": "11111111111111"},
    )
    assert resposta.status_code == 400


def test_criar_entidade_para_cpl_exige_papel_gestao_escopado(admin_client, client, db_session):
    cpl_a = _criar_cpl(admin_client, sigla="CPL-GESTORA-04A")
    cpl_b = _criar_cpl(admin_client, sigla="CPL-GESTORA-04B")
    gestor_a = criar_usuario_com_papel(db_session, Papel.ENTIDADE_GESTORA, cpl_id=cpl_a)
    client_gestor_a = login_como(client, gestor_a)

    resposta = client_gestor_a.post(
        f"/painel/cadastro/cpls/{cpl_b}/entidades",
        data={"tipo": "empresa", "razao_social": "Fora do escopo Ltda"},
    )
    assert resposta.status_code == 403


# --- A: entidade gestora + usuário responsável ---------------------------------


def test_criar_entidade_gestora_define_cpl_entidade_gestora(admin_client):
    cpl_id = _criar_cpl(admin_client, sigla="CPL-GESTORA-05")
    resposta = admin_client.post(
        f"/painel/cpls/{cpl_id}/entidade-gestora",
        data={"tipo": "orgao_publico", "razao_social": "Secretaria de Teste"},
        follow_redirects=False,
    )
    assert resposta.status_code == 303

    cpl = admin_client.get(f"/api/cpls/{cpl_id}").json()
    assert cpl["entidade_gestora_id"] is not None


def test_criar_entidade_gestora_exige_administrador(admin_client, client, db_session):
    cpl_id = _criar_cpl(admin_client, sigla="CPL-GESTORA-06")
    gestor = criar_usuario_com_papel(db_session, Papel.ENTIDADE_GESTORA, cpl_id=cpl_id)
    client_gestor = login_como(client, gestor)

    resposta = client_gestor.post(
        f"/painel/cpls/{cpl_id}/entidade-gestora",
        data={"tipo": "orgao_publico", "razao_social": "Não devia conseguir"},
    )
    assert resposta.status_code == 403


def test_criar_usuario_responsavel_sem_entidade_gestora_e_rejeitado(admin_client):
    cpl_id = _criar_cpl(admin_client, sigla="CPL-GESTORA-07")
    resposta = admin_client.post(
        f"/painel/cpls/{cpl_id}/usuario-responsavel",
        data={
            "nome": "Fulano",
            "email": "fulano-sem-gestora@teste.com",
            "password": "SenhaForte123!",
            "papel": "entidade_gestora",
        },
    )
    assert resposta.status_code == 400


def test_criar_usuario_responsavel_cria_cadeia_completa_e_permite_login(admin_client, client):
    cpl_id = _criar_cpl(admin_client, sigla="CPL-GESTORA-08")
    admin_client.post(
        f"/painel/cpls/{cpl_id}/entidade-gestora",
        data={"tipo": "orgao_publico", "razao_social": "Secretaria Responsavel"},
    )
    resposta = admin_client.post(
        f"/painel/cpls/{cpl_id}/usuario-responsavel",
        data={
            "nome": "Beltrana Responsavel",
            "email": "beltrana.responsavel@teste.com",
            "password": "SenhaForte123!",
            "papel": "dirigente_entidade_gestora",
        },
        follow_redirects=False,
    )
    assert resposta.status_code == 303

    # `client`/`admin_client` compartilham a mesma instância de TestClient —
    # logar como a nova usuária aqui sobrescreve o header dela, então esse
    # login precisa ser a última coisa feita com este client neste teste.
    login = client.post(
        "/api/auth/login",
        data={"username": "beltrana.responsavel@teste.com", "password": "SenhaForte123!"},
    )
    assert login.status_code == 200
    client.headers.update({"Authorization": f"Bearer {login.json()['access_token']}"})

    resposta_criar_entidade = client.post(
        "/api/entidades", json={"tipo": "empresa", "razao_social": "Criada pela Beltrana"}
    )
    assert resposta_criar_entidade.status_code == 201


def test_criar_usuario_responsavel_email_duplicado_e_rejeitado(admin_client):
    cpl_id = _criar_cpl(admin_client, sigla="CPL-GESTORA-09")
    admin_client.post(
        f"/painel/cpls/{cpl_id}/entidade-gestora",
        data={"tipo": "orgao_publico", "razao_social": "Secretaria Dup"},
    )
    dados = {
        "nome": "Duplicado",
        "email": "duplicado.responsavel@teste.com",
        "password": "SenhaForte123!",
        "papel": "entidade_gestora",
    }
    primeira = admin_client.post(
        f"/painel/cpls/{cpl_id}/usuario-responsavel", data=dados, follow_redirects=False
    )
    assert primeira.status_code == 303
    segunda = admin_client.post(f"/painel/cpls/{cpl_id}/usuario-responsavel", data=dados)
    assert segunda.status_code == 400


def test_criar_usuario_responsavel_papel_invalido_e_rejeitado(admin_client):
    cpl_id = _criar_cpl(admin_client, sigla="CPL-GESTORA-10")
    admin_client.post(
        f"/painel/cpls/{cpl_id}/entidade-gestora",
        data={"tipo": "orgao_publico", "razao_social": "Secretaria Papel Invalido"},
    )
    resposta = admin_client.post(
        f"/painel/cpls/{cpl_id}/usuario-responsavel",
        data={
            "nome": "Tentando Ser Admin",
            "email": "tentando.admin@teste.com",
            "password": "SenhaForte123!",
            "papel": "administrador_plataforma",
        },
    )
    assert resposta.status_code == 400
